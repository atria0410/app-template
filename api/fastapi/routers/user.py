from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from sqlalchemy.orm import Session

import crud.user
from dependencies import get_current_user, get_db
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse
from schemas import UserSchema

router = APIRouter(prefix="/users", tags=["users"])


# ユーザー一覧取得
@router.get("/", response_model=list[UserSchema.User])
async def read_users(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: UserSchema.User = Depends(get_current_user),
):
    db_users = crud.user.get_users(db, skip=skip, limit=limit)
    return db_users


# ユーザー取得
@router.get("/{user_id}", response_model=UserSchema.User)
async def read_user(
    user_id: int, db: Session = Depends(get_db), current_user: UserSchema.User = Depends(get_current_user)
):
    db_user = crud.user.get_user(db, user_id=user_id)
    if not db_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ユーザーが存在しません。")
    return db_user


# ユーザー登録
@router.post("/", response_model=UserSchema.User)
async def create_user(
    user: UserSchema.UserCreate,
    db: Session = Depends(get_db),
    current_user: UserSchema.User = Depends(get_current_user),
):
    db_user = crud.user.get_user_by_email(db, email=user.email)
    if db_user:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="メールアドレスが既に利用されています。")
    return crud.user.create_user(db=db, user=user)


# ユーザー更新
@router.put("/{user_id}", response_model=UserSchema.User)
async def update_user(
    user_id: int,
    user: UserSchema.UserUpdate,
    db: Session = Depends(get_db),
    current_user: UserSchema.User = Depends(get_current_user),
):
    db_user = crud.user.update_user(db, user_id=user_id, user=user)
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ユーザーが存在しません。")
    return db_user


# ユーザー削除
@router.delete("/", response_model=UserSchema.User)
async def delete_user(
    user_id: int, db: Session = Depends(get_db), current_user: UserSchema.User = Depends(get_current_user)
):
    db_user = crud.user.delete_user(db, user_id=user_id)
    if not db_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="ユーザーが存在しません。")
    return db_user


# ユーザー一覧帳票取得
@router.get("/download/xlsx", response_model=list[UserSchema.User])
async def read_users(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: UserSchema.User = Depends(get_current_user),
):
    db_users = crud.user.get_users(db, skip=skip, limit=limit)

    template_path = "report/user_list.xlsx"
    save_path = "tmp/user_list.xlsx"

    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side)

    wb = load_workbook(template_path)
    ws = wb["Sheet1"]
    for row, user in enumerate(db_users, start=2):
        ws.cell(row=row, column=1, value=user.id).border = border
        ws.cell(row=row, column=2, value=user.email).border = border
        ws.cell(row=row, column=3, value=user.first_name).border = border
        ws.cell(row=row, column=4, value=user.last_name).border = border
        ws.cell(row=row, column=5, value=user.birthday).border = border
        ws.cell(row=row, column=6, value=user.is_active).border = border
    wb.save(save_path)

    return FileResponse(
        path=save_path,
        headers={"Content-Disposition": 'attachment; filename="user_list.xlsx"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
